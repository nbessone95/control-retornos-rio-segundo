import streamlit as st
from datetime import datetime
import os
import pandas as pd
from PIL import Image
from openpyxl import load_workbook
from streamlit_drawable_canvas import st_canvas   # ← Esta línea faltaba

st.set_page_config(page_title="Control Retornos", layout="wide")
st.title("🧾 Control de Retornos - Rio Segundo")

os.makedirs("templates", exist_ok=True)
os.makedirs("completed", exist_ok=True)
os.makedirs("data", exist_ok=True)

menu = st.sidebar.selectbox("Menú", ["Completar Formulario", "Subir Plantillas", "Historial"])

if menu == "Subir Plantillas":
    st.header("📤 Subir Plantillas")
    uploaded = st.file_uploader("Subir Excel", type="xlsx", accept_multiple_files=True)
    if uploaded:
        for file in uploaded:
            with open(f"templates/{file.name}", "wb") as f:
                f.write(file.getbuffer())
            st.success(f"✅ {file.name} guardado")
        st.rerun()

elif menu == "Completar Formulario":
    st.header("✍️ Completar Control")
    
    files = [f for f in os.listdir("templates") if f.endswith(".xlsx")]
    if not files:
        st.warning("Sube una plantilla primero")
    else:
        selected = st.selectbox("Seleccionar archivo", files)
        
        if st.button("Cargar Formulario", type="primary"):
            st.session_state.selected_file = selected
            st.rerun()
        
        if 'selected_file' in st.session_state:
            filepath = f"templates/{st.session_state.selected_file}"
            st.title(f"🧾 {st.session_state.selected_file.replace('.xlsx', '')}")
            
            localidad = "Rio Segundo"
            equipo = "Alessandrini / Rosso / Baldoncini"
            total_desp = 0.0
            try:
                wb = load_workbook(filepath, data_only=True)
                ws3 = wb["Hoja3"] if "Hoja3" in wb.sheetnames else wb.active
                localidad = str(ws3.cell(row=2, column=2).value or localidad)
                if "Hoja2" in wb.sheetnames:
                    ws2 = wb["Hoja2"]
                    equipo = str(ws2.cell(row=2, column=2).value or equipo)
                total_desp = float(ws3.cell(row=24, column=5).value or 0)
            except:
                pass

            st.success(f"Trabajando con: **{st.session_state.selected_file}**")

            # Datos Solo Lectura
            st.subheader("1. Datos Generales (no modificables)")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Localidad", localidad)
                st.metric("Equipo", equipo)
                st.metric("Camión", "AD")
            with col2:
                st.metric("Fecha", datetime.today().strftime("%d-%m-%Y"))
                st.metric("Clientes", 17)
                st.metric("Total Despachado", total_desp)

            # Formulario
            st.subheader("2. Retornos y Cambios")
            c1, c2 = st.columns(2)
            with c1:
                ret_2500 = st.number_input("Retorno 2500", value=0.0, step=0.01, format="%.2f")
                ret_2000 = st.number_input("Retorno 2000", value=0.0, step=0.01, format="%.2f")
            with c2:
                ret_1250 = st.number_input("Retorno 1250", value=0.0, step=0.01, format="%.2f")

            st.subheader("Retornos Llenos")
            rl1, rl2 = st.columns(2)
            with rl1:
                lleno_2500 = st.number_input("Lleno 2500", value=0.0, step=0.01, format="%.2f")
                lleno_2000 = st.number_input("Lleno 2000", value=0.0, step=0.01, format="%.2f")
            with rl2:
                lleno_1250 = st.number_input("Lleno 1250", value=0.0, step=0.01, format="%.2f")

            st.subheader("Cambios")
            c1, c2, c3 = st.columns(3)
            with c1:
                cam_2500 = st.number_input("Cambio 2500", value=0.0, step=0.01, format="%.2f")
                cam_2000 = st.number_input("Cambio 2000", value=0.0, step=0.01, format="%.2f")
            with c2:
                cam_1250 = st.number_input("Cambio 1250", value=0.0, step=0.01, format="%.2f")
                cam_354 = st.number_input("Cambio 354", value=0.0, step=0.01, format="%.2f")
            with c3:
                cam_220 = st.number_input("Cambio 220", value=0.0, step=0.01, format="%.2f")
                cam_473 = st.number_input("Cambio 473", value=0.0, step=0.01, format="%.2f")

            st.subheader("Otras Operaciones")
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                venta_envases = st.number_input("Venta de Envases", value=0.0, step=0.01, format="%.2f")
            with col_b:
                prestamos = st.number_input("Préstamos", value=0.0, step=0.01, format="%.2f")
            with col_c:
                retiros = st.number_input("Retiros", value=0.0, step=0.01, format="%.2f")

            observaciones = st.text_area("Observaciones", height=80)

            # ==================== FIRMAS DIGITALES ====================
            st.subheader("✍️ Firmas Digitales")
            col_f1, col_f2 = st.columns(2)
            
            with col_f1:
                st.write("**Firma del Repartidor**")
                canvas1 = st_canvas(
                    height=180,
                    width=350,
                    stroke_width=3,
                    stroke_color="#000000",
                    background_color="#ffffff",
                    key="canvas_repartidor"
                )
            
            with col_f2:
                st.write("**Firma del Controlador**")
                canvas2 = st_canvas(
                    height=180,
                    width=350,
                    stroke_width=3,
                    stroke_color="#000000",
                    background_color="#ffffff",
                    key="canvas_controlador"
                )

            if st.button("💾 Guardar Control Completo", type="primary"):
                if canvas1.image_data is None or canvas2.image_data is None:
                    st.error("Por favor realice ambas firmas digitales")
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                    
                    img1 = Image.fromarray(canvas1.image_data.astype("uint8"))
                    img2 = Image.fromarray(canvas2.image_data.astype("uint8"))
                    firma1_path = f"completed/firma_repartidor_{timestamp}.png"
                    firma2_path = f"completed/firma_controlador_{timestamp}.png"
                    img1.save(firma1_path)
                    img2.save(firma2_path)
                    
                    data = {
                        "Fecha": datetime.today().strftime("%d-%m-%Y"),
                        "Hora": datetime.today().strftime("%H:%M"),
                        "Localidad": localidad,
                        "Equipo": equipo,
                        "Camion": "AD",
                        "Archivo": st.session_state.selected_file,
                        "Total_Despachado": total_desp,
                        "Retorno_2500": ret_2500,
                        "Retorno_2000": ret_2000,
                        "Retorno_1250": ret_1250,
                        "Lleno_2500": lleno_2500,
                        "Lleno_2000": lleno_2000,
                        "Lleno_1250": lleno_1250,
                        "Cambio_2500": cam_2500,
                        "Cambio_2000": cam_2000,
                        "Cambio_1250": cam_1250,
                        "Cambio_354": cam_354,
                        "Cambio_220": cam_220,
                        "Cambio_473": cam_473,
                        "Venta_Envases": venta_envases,
                        "Prestamos": prestamos,
                        "Retiros": retiros,
                        "Observaciones": observaciones,
                        "Firma_Repartidor": firma1_path,
                        "Firma_Controlador": firma2_path
                    }
                    pd.DataFrame([data]).to_csv(f"data/control_{timestamp}.csv", index=False)

                    st.success("✅ ¡Control guardado correctamente!")
                    st.image(firma1_path, caption="Firma Repartidor")
                    st.image(firma2_path, caption="Firma Controlador")
                    st.balloons()

else:
    st.header("📋 Historial de Controles")
    data_files = [f for f in os.listdir("data") if f.endswith(".csv")]
    if data_files:
        for f in sorted(data_files, reverse=True):
            df = pd.read_csv(f"data/{f}")
            st.subheader(f"📅 {df['Fecha'].iloc[0]} {df['Hora'].iloc[0]} - {df['Archivo'].iloc[0]}")
            st.caption(f"Camión: {df['Camion'].iloc[0]} | Total Despachado: {df['Total_Despachado'].iloc[0]}")
            st.dataframe(df, use_container_width=True)
            
            if "Firma_Repartidor" in df.columns and os.path.exists(df["Firma_Repartidor"].iloc[0]):
                st.image(df["Firma_Repartidor"].iloc[0], caption="Firma Repartidor")
            if "Firma_Controlador" in df.columns and os.path.exists(df["Firma_Controlador"].iloc[0]):
                st.image(df["Firma_Controlador"].iloc[0], caption="Firma Controlador")
            
            st.divider()
    else:
        st.info("Aún no hay controles guardados.")
